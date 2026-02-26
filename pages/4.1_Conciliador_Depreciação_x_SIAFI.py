import streamlit as st
import pandas as pd
import pdfplumber
import re
from fpdf import FPDF, XPos, YPos
import io
import os
import openpyxl
import copy

# ==========================================
# CONFIGURAÇÃO INICIAL E MEMÓRIA
# ==========================================
st.set_page_config(
    page_title="Conciliador de Depreciação",
    page_icon="📊",
    layout="wide"
)

# Inicializa a memória do Streamlit
if 'dados_processados' not in st.session_state:
    st.session_state.dados_processados = False
if 'dados_ug' not in st.session_state:
    st.session_state.dados_ug = {}
if 'matriz_erro' not in st.session_state:
    st.session_state.matriz_erro = False

# Oculta elementos técnicos da plataforma e o menu lateral padrão
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            [data-testid="stSidebar"] {display: none;}
            [data-testid="collapsedControl"] {display: none;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# Botão para retornar à tela inicial solto no topo da tela
st.page_link("Menu_principal.py", label="⬅️ Voltar ao Menu Inicial")

# ==========================================
# FUNÇÕES DE PROCESSAMENTO (BASTIDORES)
# ==========================================
def formatar_real(valor):
    sinal = "-" if valor < -0.001 else ""
    return f"{sinal}{abs(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_moeda_pdf(valor_str):
    if not valor_str: return 0.0
    try:
        if len(valor_str) >= 3 and valor_str[-3] in [',', '.']:
            inteiro = valor_str[:-3].replace('.', '').replace(',', '')
            decimal = valor_str[-2:]
            return float(f"{inteiro}.{decimal}")
        else:
            limpo = valor_str.replace('.', '').replace(',', '.')
            return float(limpo)
    except:
        return 0.0

def converter_valor_excel(valor):
    if pd.isna(valor) or valor is None: return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    v_str = str(valor).strip().replace('R$', '').replace(' ', '')
    if ',' in v_str: v_str = v_str.replace('.', '').replace(',', '.')
    try: return float(v_str)
    except: return 0.0

def extrair_codigo_grupo(valor_nat_desp):
    try:
        if isinstance(valor_nat_desp, float): valor_nat_desp = int(valor_nat_desp)
        s_val = re.sub(r'\D', '', str(valor_nat_desp).strip())
        if len(s_val) < 5: return None
        return int(s_val[-2:])
    except: return None

def extrair_id_unidade(texto):
    match = re.search(r"(\d+)", str(texto))
    return match.group(1) if match else None

def extrair_valor_mes(bloco_texto, nome_linha, idx_mes):
    nome_regex = nome_linha.replace("(", r"\(").replace(")", r"\)").replace(" ", r"\s+")
    nome_regex = nome_regex.replace("Ç", "[CÇ]").replace("Ã", "[AÃ]").replace("Ê", "[EÊ]").replace("Í", "[IÍ]")
    
    proximos_rotulos = r"(?:SALDO INICIAL|DEPRECIA[CÇ][AÃ]O M[EÊ]S|ENTRADAS|SA[IÍ]DAS|AJUSTE|TOTAL|SALDO ATUAL|\(\*\)|$)"
    
    regex = re.compile(rf"{nome_regex}([\s\S]*?)(?={proximos_rotulos})", re.IGNORECASE)
    match = regex.search(bloco_texto)
    if match:
        trecho = match.group(1)
        padrao = re.compile(r"\d{1,3}(?:[.,]\d{3})*[.,]\d{2}")
        matches = padrao.findall(trecho)
        if len(matches) > idx_mes:
            return formatar_moeda_pdf(matches[idx_mes])
    return 0.0

def processar_pdf(arquivo_obj, idx_mes):
    dados_pdf = {}
    texto_completo = ""
    try:
        with pdfplumber.open(arquivo_obj) as pdf:
            for page in pdf.pages: texto_completo += page.extract_text() + "\n"
    except Exception:
        return {}

    regex_cabecalho = re.compile(r"(?m)^\s*(\d+)\s*-\s*[A-Z]")
    matches = list(regex_cabecalho.finditer(texto_completo))
    
    for i, match in enumerate(matches):
        grupo_id = int(match.group(1))
        start_idx = match.start()
        end_idx = matches[i+1].start() if i + 1 < len(matches) else len(texto_completo)
        bloco_texto = texto_completo[start_idx:end_idx]
        
        regex_saldo = re.compile(r"\(\*\)\s*SALDO[\s\S]*?ATUAL[\s\S]*?((?:\d{1,3}(?:\.\d{3})*,\d{2}))")
        match_saldo = regex_saldo.search(bloco_texto)
        saldo_val = formatar_moeda_pdf(match_saldo.group(1)) if match_saldo else 0.0
        
        v_dep_mes = extrair_valor_mes(bloco_texto, "DEPRECIAÇÃO MÊS CORRENTE", idx_mes)
        v_entradas = extrair_valor_mes(bloco_texto, "ENTRADAS (TRANSFERÊNCIA)", idx_mes)
        v_saidas_transf = extrair_valor_mes(bloco_texto, "SAÍDAS (TRANSFERÊNCIA)", idx_mes)
        v_saidas_baixas = extrair_valor_mes(bloco_texto, "SAÍDAS (BAIXAS)", idx_mes)
        
        movim_val = v_dep_mes + v_entradas - v_saidas_transf - v_saidas_baixas
            
        dados_pdf[grupo_id] = {'saldo': saldo_val, 'movimento': movim_val}
            
    return dados_pdf

def carregar_matriz():
    caminho_matriz = "MATRIZ.xlsx"
    dicionario_matriz = {}
    if os.path.exists(caminho_matriz):
        wb_matriz = openpyxl.load_workbook(caminho_matriz, data_only=True)
        ws_matriz = wb_matriz.active
        for row in ws_matriz.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] is not None:
                dicionario_matriz[str(row[0]).strip()] = str(row[1]).strip()
    return dicionario_matriz

class PDFRelatorio(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 12)
        self.cell(0, 10, 'Relatório de Conciliação - Depreciação Acumulada', align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        self.ln(5)
        
    def footer(self):
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', align='C')

# ==========================================
# INTERFACE DO USUÁRIO
# ==========================================
st.title("📊 Conciliador de Depreciação")

meses_opcoes = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
mes_selecionado = st.selectbox("Selecione o Mês de Referência:", meses_opcoes)
idx_mes = meses_opcoes.index(mes_selecionado)

with st.expander("📘 GUIA DE USO (Clique para abrir)", expanded=False):
    st.markdown("📌 **Orientações de Uso**")
    st.markdown("""
    1. **Selecione o Mês** que deseja conferir acima.
    2. Anexe a **Planilha do tesouro e os relatórios** todos juntos no mesmo local abaixo.
    3. O sistema fará a leitura inicial. **Poderá corrigir valores divergentes manualmente e as edições serão registadas no PDF!**
    """)

uploaded_files = st.file_uploader(
    "📂 Arraste ou selecione a Planilha e os Relatórios de uma só vez", 
    type=['pdf', 'xlsx'], 
    accept_multiple_files=True
)

# ==========================================
# ETAPA 1: PROCESSAMENTO DE DADOS
# ==========================================
if st.button("🚀 Gerar Relatório de Conciliação", type="primary", use_container_width=True):
    if not uploaded_files:
        st.warning("⚠️ Por favor, insira seus arquivos para que possamos realizar a conciliação.")
    else:
        pdfs = [f for f in uploaded_files if f.name.lower().endswith('.pdf')]
        excel_alvo = [f for f in uploaded_files if f.name.lower().endswith('.xlsx')]
        
        if not excel_alvo:
            st.error("❌ A planilha base em Excel não foi encontrada junto com os arquivos enviados.")
            st.stop()
            
        arquivo_alvo = excel_alvo[0]

        with st.spinner("Analisando seus documentos e cruzando os valores..."):
            try:
                dicionario_matriz = carregar_matriz()
                if not dicionario_matriz:
                    st.error("❌ Matriz não encontrada ou vazia. Verifique o arquivo MATRIZ.xlsx.")
                    st.stop()

                dados_pdfs_extraidos = {}
                for f in pdfs:
                    uid = extrair_id_unidade(f.name)
                    if uid: dados_pdfs_extraidos[uid] = processar_pdf(f, idx_mes)

                wb_alvo = openpyxl.load_workbook(arquivo_alvo, data_only=True)
                abas = [s for s in wb_alvo.sheetnames if s != "MATRIZ"]
                
                progresso = st.progress(0)
                status_box = st.empty()
                
                dados_ug = {}

                for idx, sheet_name in enumerate(abas):
                    ws = wb_alvo[sheet_name]
                    uid = extrair_id_unidade(sheet_name)
                    status_box.text(f"Lendo e analisando dados da Unidade Gestora: {sheet_name}...")
                    
                    d_excel = {}
                    for row in ws.iter_rows(values_only=True):
                        if not row or not row[0]: continue
                        conta_raw = str(row[0]).strip()
                        if conta_raw.startswith("12") and conta_raw.replace('.', '').isdigit():
                            if conta_raw == "123110402": continue
                            nat_desp = dicionario_matriz.get(conta_raw)
                            if nat_desp:
                                grupo = extrair_codigo_grupo(nat_desp)
                                if grupo is not None:
                                    valid_vals = [v for v in row if v is not None and str(v).strip() != ""]
                                    if len(valid_vals) >= 2:
                                        saldo_raw, movim_raw = valid_vals[-1], valid_vals[-2]
                                    elif len(valid_vals) == 1:
                                        saldo_raw, movim_raw = valid_vals[-1], 0.0
                                    else:
                                        saldo_raw, movim_raw = 0.0, 0.0
                                            
                                    val_saldo = converter_valor_excel(saldo_raw)
                                    val_mov = converter_valor_excel(movim_raw)
                                    
                                    if grupo not in d_excel: d_excel[grupo] = {'saldo': 0.0, 'movimento': 0.0}
                                    d_excel[grupo]['saldo'] += val_saldo
                                    d_excel[grupo]['movimento'] += val_mov

                    # Prepara a comparação com o PDF
                    d_pdf_raw = dados_pdfs_extraidos.get(uid, {})
                    grupos_combinados = sorted(list(set(d_pdf_raw.keys()) | set(d_excel.keys())))
                    
                    d_pdf = {}
                    grupos_com_erro = []
                    erro_original = False
                    
                    for g in grupos_combinados:
                        # Extrai e converte para o sinal correto para a conciliação (inversão)
                        vp_saldo = round(-1 * d_pdf_raw.get(g, {}).get('saldo', 0.0), 2)
                        vp_mov = round(-1 * d_pdf_raw.get(g, {}).get('movimento', 0.0), 2)
                        
                        ve_saldo = round(d_excel.get(g, {}).get('saldo', 0.0), 2)
                        ve_mov = round(d_excel.get(g, {}).get('movimento', 0.0), 2)
                        
                        d_pdf[g] = {'saldo': vp_saldo, 'movimento': vp_mov}
                        
                        if abs(vp_saldo - ve_saldo) > 0.00 or abs(vp_mov - ve_mov) > 0.00:
                            erro_original = True
                            grupos_com_erro.append(g)

                    dados_ug[sheet_name] = {
                        'uid': uid,
                        'd_excel': d_excel,
                        'd_pdf': d_pdf,
                        'd_pdf_orig': copy.deepcopy(d_pdf), # Memória fotográfica
                        'tem_pdf': uid in dados_pdfs_extraidos,
                        'erro_original': erro_original,
                        'grupos_com_erro': grupos_com_erro
                    }

                    progresso.progress((idx + 1) / len(abas))

                st.session_state.dados_ug = dados_ug
                st.session_state.dados_processados = True
                progresso.empty()
                status_box.empty()
                
            except Exception as e:
                st.error(f"Não foi possível processar seus documentos. Verifique os ficheiros. (Detalhe: {e})")

# ==========================================
# ETAPA 2: REVISÃO CIRÚRGICA E PDF FINAL
# ==========================================
if st.session_state.get('dados_processados'):
    st.markdown("---")
    st.subheader("🔍 Resultados da Conciliação & Revisão")
    st.info("💡 **Ação Cirúrgica:** Apenas os grupos com divergências (ou ficheiros em falta) exibem campos de edição. As edições ficarão registadas no PDF Final.")

    pdf_out = PDFRelatorio()
    pdf_out.set_auto_page_break(auto=True, margin=15)
    pdf_out.add_page()
    
    lista_resumo = []
    dados_ug = st.session_state.dados_ug

    for sheet_name, info in dados_ug.items():
        uid = info['uid']
        d_excel = info['d_excel']
        d_pdf = info['d_pdf']
        d_pdf_orig = info['d_pdf_orig']
        
        # 1. ATUALIZA VALORES EM TEMPO REAL
        if info['erro_original'] or not info['tem_pdf']:
            for g in info['grupos_com_erro']:
                k_saldo = f"ed_s_{sheet_name}_{g}"
                k_mov = f"ed_m_{sheet_name}_{g}"
                if k_saldo in st.session_state: d_pdf[g]['saldo'] = st.session_state[k_saldo]
                if k_mov in st.session_state: d_pdf[g]['movimento'] = st.session_state[k_mov]

        # 2. RECÁLCULO
        divergencias = []
        alertas_auditoria = []
        soma_pdf_s = soma_excel_s = soma_pdf_m = soma_excel_m = 0.0
        
        grupos = sorted(list(d_pdf.keys()))
        for g in grupos:
            vp_s = d_pdf[g]['saldo']
            ve_s = d_excel.get(g, {}).get('saldo', 0.0)
            vp_m = d_pdf[g]['movimento']
            ve_m = d_excel.get(g, {}).get('movimento', 0.0)
            
            soma_pdf_s += vp_s
            soma_excel_s += ve_s
            soma_pdf_m += vp_m
            soma_excel_m += ve_m
            
            dif_s = round(vp_s - ve_s, 2)
            if abs(dif_s) > 0.00:
                divergencias.append({'grupo': g, 'tipo': 'Saldo Acumulado', 'pdf': vp_s, 'excel': ve_s, 'diff': dif_s})
            
            dif_m = round(vp_m - ve_m, 2)
            if abs(dif_m) > 0.00:
                divergencias.append({'grupo': g, 'tipo': 'Mês Corrente', 'pdf': vp_m, 'excel': ve_m, 'diff': dif_m})
            
            # Auditoria por grupo
            if abs(vp_s - d_pdf_orig[g]['saldo']) > 0.01:
                alertas_auditoria.append(f"* ALERTA: O Saldo Acumulado do Grupo {g} foi alterado manualmente pelo utilizador. (Original lido do PDF: R$ {formatar_real(d_pdf_orig[g]['saldo'])})")
            if abs(vp_m - d_pdf_orig[g]['movimento']) > 0.01:
                alertas_auditoria.append(f"* ALERTA: O Mês Corrente do Grupo {g} foi alterado manualmente pelo utilizador. (Original lido do PDF: R$ {formatar_real(d_pdf_orig[g]['movimento'])})")

        dif_total_saldo = round(soma_pdf_s - soma_excel_s, 2)
        dif_total_mov = round(soma_pdf_m - soma_excel_m, 2)
        tem_erro_atual = len(divergencias) > 0

        # 3. EXIBIÇÃO EM TELA
        with st.container():
            st.markdown(f"### 🏢 Unidade Gestora: {sheet_name} (ID: {uid})")
            
            col1, col2 = st.columns(2)
            col1.metric("Diferença Saldo Acumulado", f"R$ {formatar_real(dif_total_saldo)}", delta_color="inverse" if abs(dif_total_saldo) > 0.05 else "normal")
            col2.metric("Diferença Mês Corrente", f"R$ {formatar_real(dif_total_mov)}", delta_color="inverse" if abs(dif_total_mov) > 0.05 else "normal")
            
            titulo_expander = "⚠️ Grupos com Divergência" if tem_erro_atual else ("✅ Corrigido Manualmente" if info['erro_original'] else "✅ Conciliado")
            if not info['tem_pdf']: titulo_expander += " (Relatório Ausente)"
            
            with st.expander(titulo_expander, expanded=tem_erro_atual or not info['tem_pdf']):
                if divergencias:
                    df_view = pd.DataFrame([{
                        'Grupo': d['grupo'],
                        'Tipo': d['tipo'],
                        'Valor Relatório': d['pdf'],
                        'Valor Planilha': d['excel'],
                        'Diferença': d['diff']
                    } for d in divergencias])
                    
                    st.dataframe(df_view.style.format({
                        "Valor Relatório": lambda x: f"R$ {formatar_real(x)}",
                        "Valor Planilha": lambda x: f"R$ {formatar_real(x)}",
                        "Diferença": lambda x: f"R$ {formatar_real(x)}"
                    }), use_container_width=True)
                else:
                    st.success("Nenhuma divergência nesta unidade.")

                # Caixas de Edição (Aparecem apenas nos grupos com erro na extração original ou relatórios ausentes)
                if info['grupos_com_erro']:
                    st.markdown("---")
                    st.markdown("**✏️ Correção Direta por Grupo Divergente:**")
                    for g in info['grupos_com_erro']:
                        st.markdown(f"**🔹 Grupo {g}**")
                        c1, c2 = st.columns(2)
                        with c1:
                            st.number_input(f"Saldo Acumulado (Relatório)", value=float(d_pdf[g]['saldo']), step=100.0, key=f"ed_s_{sheet_name}_{g}")
                        with c2:
                            st.number_input(f"Mês Corrente (Relatório)", value=float(d_pdf[g]['movimento']), step=100.0, key=f"ed_m_{sheet_name}_{g}")

        # 4. ESCRITA NO PDF FINAL
        if pdf_out.get_y() > 240: pdf_out.add_page()
        
        pdf_out.set_font("helvetica", 'B', 10)
        pdf_out.set_fill_color(240, 240, 240)
        pdf_out.cell(0, 8, f"Unidade Gestora: {sheet_name} (ID: {uid})", 1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf_out.ln(2)
        
        pdf_out.set_font("helvetica", 'B', 9)
        pdf_out.set_fill_color(220, 230, 241)
        pdf_out.cell(48, 7, "Métrica", 1, fill=True, align='C')
        pdf_out.cell(48, 7, "Total Relatório", 1, fill=True, align='C')
        pdf_out.cell(48, 7, "Total Planilha", 1, fill=True, align='C')
        pdf_out.cell(46, 7, "Diferença", 1, fill=True, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf_out.set_font("helvetica", '', 9)
        
        edit_geral_saldo = any(abs(d_pdf[g]['saldo'] - d_pdf_orig[g]['saldo']) > 0.01 for g in grupos)
        edit_geral_mov = any(abs(d_pdf[g]['movimento'] - d_pdf_orig[g]['movimento']) > 0.01 for g in grupos)
        
        str_pdf_s = f"R$ {formatar_real(soma_pdf_s)}" + (" *" if edit_geral_saldo else "")
        str_pdf_m = f"R$ {formatar_real(soma_pdf_m)}" + (" *" if edit_geral_mov else "")

        # Resumo Métrica Saldo
        pdf_out.cell(48, 7, "Saldo Acumulado", 1)
        pdf_out.cell(48, 7, str_pdf_s, 1, align='R')
        pdf_out.cell(48, 7, f"R$ {formatar_real(soma_excel_s)}", 1, align='R')
        if abs(dif_total_saldo) > 0.00: pdf_out.set_text_color(200, 0, 0)
        pdf_out.cell(46, 7, f"R$ {formatar_real(dif_total_saldo)}", 1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf_out.set_text_color(0, 0, 0)
        
        # Resumo Métrica Movimento
        pdf_out.cell(48, 7, f"Mês Corrente ({mes_selecionado})", 1)
        pdf_out.cell(48, 7, str_pdf_m, 1, align='R')
        pdf_out.cell(48, 7, f"R$ {formatar_real(soma_excel_m)}", 1, align='R')
        if abs(dif_total_mov) > 0.00: pdf_out.set_text_color(200, 0, 0)
        pdf_out.cell(46, 7, f"R$ {formatar_real(dif_total_mov)}", 1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf_out.set_text_color(0, 0, 0)
        
        pdf_out.ln(3)
        
        status_str = "✅ Conciliado"
        if not divergencias:
            pdf_out.set_fill_color(220, 255, 220)
            pdf_out.set_font("helvetica", 'B', 9)
            pdf_out.cell(0, 8, "CONCILIADO - SEM DIVERGÊNCIAS", 1, fill=True, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        else:
            status_str = f"❌ Divergência(s)"
            pdf_out.set_fill_color(255, 220, 220)
            pdf_out.set_font("helvetica", 'B', 9)
            pdf_out.cell(0, 8, "DIVERGÊNCIAS ENCONTRADAS:", 1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            pdf_out.set_fill_color(250, 250, 250)
            pdf_out.set_font("helvetica", 'B', 8)
            pdf_out.cell(15, 6, "Grupo", 1, fill=True, align='C')
            pdf_out.cell(35, 6, "Tipo", 1, fill=True, align='C')
            pdf_out.cell(46, 6, "Valor Relatório", 1, fill=True, align='C')
            pdf_out.cell(46, 6, "Valor Planilha", 1, fill=True, align='C')
            pdf_out.cell(48, 6, "Diferença", 1, fill=True, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            
            pdf_out.set_font("helvetica", '', 8)
            for d in divergencias:
                g = d['grupo']
                if d['tipo'] == 'Saldo Acumulado': is_edit = abs(d_pdf[g]['saldo'] - d_pdf_orig[g]['saldo']) > 0.01
                else: is_edit = abs(d_pdf[g]['movimento'] - d_pdf_orig[g]['movimento']) > 0.01
                    
                val_pdf_str = f"R$ {formatar_real(d['pdf'])}" + (" *" if is_edit else "")
                
                pdf_out.cell(15, 6, str(d['grupo']), 1, align='C')
                pdf_out.cell(35, 6, d['tipo'], 1, align='C')
                pdf_out.cell(46, 6, val_pdf_str, 1, align='R')
                pdf_out.cell(46, 6, f"R$ {formatar_real(d['excel'])}", 1, align='R')
                pdf_out.set_text_color(200, 0, 0)
                pdf_out.cell(48, 6, f"R$ {formatar_real(d['diff'])}", 1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf_out.set_text_color(0, 0, 0)
        
        # LOGS DE AUDITORIA NO PDF
        if alertas_auditoria:
            pdf_out.set_font("helvetica", 'I', 7)
            pdf_out.set_text_color(180, 0, 0)
            for alerta in alertas_auditoria:
                pdf_out.cell(0, 5, alerta, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf_out.set_text_color(0, 0, 0)
                
        pdf_out.ln(5)
        pdf_out.cell(0, 0, "", "B", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf_out.ln(5)
        
        if not info['tem_pdf']: status_str = "⚠️ Relatório ausente"
            
        lista_resumo.append({
            "Unidade / Aba": sheet_name,
            "Status": status_str,
            "Dif. Saldo Total": f"R$ {formatar_real(dif_total_saldo)}",
            "Dif. Mês Corrente": f"R$ {formatar_real(dif_total_mov)}"
        })

    # Resumo Final e Download
    st.markdown("### Resumo Geral da Conciliação (Atualizado em Tempo Real)")
    st.dataframe(pd.DataFrame(lista_resumo), use_container_width=True)
    
    try:
        pdf_bytes = bytes(pdf_out.output()) 
        st.download_button(
            label="📄 Fazer Download do Relatório Completo (PDF)",
            data=pdf_bytes,
            file_name="Relatorio_Conciliacao_Depreciacao.pdf",
            mime="application/pdf",
            type="primary",
            use_container_width=True
        )
    except Exception as e:
        st.error(f"Erro ao gerar o PDF para download. (Detalhe: {e})")
