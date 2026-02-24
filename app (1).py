import streamlit as st
import pandas as pd
import pdfplumber
import re
from fpdf import FPDF, XPos, YPos
import io
import os
import openpyxl

# O sistema prepara a tela inicial para que você tenha a melhor visualização
st.set_page_config(
    page_title="Conciliador de Depreciação",
    page_icon="📊",
    layout="wide"
)

# Oculta elementos técnicos da plataforma para deixar a interface mais limpa para o usuário
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            header {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)

# O sistema formata os valores para o padrão financeiro brasileiro (R$)
def formatar_real(valor):
    sinal = "-" if valor < -0.001 else ""
    return f"{sinal}{abs(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_moeda_pdf(valor_str):
    if not valor_str: return 0.0
    try:
        if len(valor_str) >= 3 and valor_str[-3] in [',', '.']:
            inteiro = valor_str[:-3].replace('.', '').replace(',', '')
            decimal = valor_str[-2:]
            return float(f"{inteiro}.{decimal}")
        else:
            limpo = valor_str.replace('.', '').replace(',', '.')
            return float(limpo)
    except:
        return 0.0

# O sistema garante a leitura precisa dos valores positivos e negativos da sua planilha
def converter_valor_excel(valor):
    if pd.isna(valor) or valor is None: return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    v_str = str(valor).strip().replace('R$', '').replace(' ', '')
    if ',' in v_str: v_str = v_str.replace('.', '').replace(',', '.')
    try: return float(v_str)
    except: return 0.0

def extrair_codigo_grupo(valor_nat_desp):
    try:
        if isinstance(valor_nat_desp, float): valor_nat_desp = int(valor_nat_desp)
        s_val = re.sub(r'\D', '', str(valor_nat_desp).strip())
        if len(s_val) < 5: return None
        return int(s_val[-2:])
    except: return None

def extrair_id_unidade(texto):
    match = re.search(r"(\d+)", str(texto))
    return match.group(1) if match else None

# O sistema varre os relatórios em busca das linhas contábeis que formam a sua movimentação
def extrair_valor_mes(bloco_texto, nome_linha, idx_mes):
    nome_regex = nome_linha.replace("(", r"\(").replace(")", r"\)").replace(" ", r"\s+")
    nome_regex = nome_regex.replace("Ç", "[CÇ]").replace("Ã", "[AÃ]").replace("Ê", "[EÊ]").replace("Í", "[IÍ]")
    
    proximos_rotulos = r"(?:SALDO INICIAL|DEPRECIA[CÇ][AÃ]O M[EÊ]S|ENTRADAS|SA[IÍ]DAS|AJUSTE|TOTAL|SALDO ATUAL|\(\*\)|$)"
    
    regex = re.compile(rf"{nome_regex}([\s\S]*?)(?={proximos_rotulos})", re.IGNORECASE)
    match = regex.search(bloco_texto)
    if match:
        trecho = match.group(1)
        padrao = re.compile(r"\d{1,3}(?:[.,]\d{3})*[.,]\d{2}")
        matches = padrao.findall(trecho)
        if len(matches) > idx_mes:
            return formatar_moeda_pdf(matches[idx_mes])
    return 0.0

# O sistema calcula o valor líquido exato cruzando as entradas, saídas e a depreciação do seu relatório
def processar_pdf(arquivo_obj, idx_mes):
    dados_pdf = {}
    texto_completo = ""
    try:
        with pdfplumber.open(arquivo_obj) as pdf:
            for page in pdf.pages: texto_completo += page.extract_text() + "\n"
    except Exception:
        return {}

    regex_cabecalho = re.compile(r"(?m)^\s*(\d+)\s*-\s*[A-Z]")
    matches = list(regex_cabecalho.finditer(texto_completo))
    
    for i, match in enumerate(matches):
        grupo_id = int(match.group(1))
        start_idx = match.start()
        end_idx = matches[i+1].start() if i + 1 < len(matches) else len(texto_completo)
        bloco_texto = texto_completo[start_idx:end_idx]
        
        regex_saldo = re.compile(r"\(\*\)\s*SALDO[\s\S]*?ATUAL[\s\S]*?((?:\d{1,3}(?:\.\d{3})*,\d{2}))")
        match_saldo = regex_saldo.search(bloco_texto)
        saldo_val = formatar_moeda_pdf(match_saldo.group(1)) if match_saldo else 0.0
        
        v_dep_mes = extrair_valor_mes(bloco_texto, "DEPRECIAÇÃO MÊS CORRENTE", idx_mes)
        v_entradas = extrair_valor_mes(bloco_texto, "ENTRADAS (TRANSFERÊNCIA)", idx_mes)
        v_saidas_transf = extrair_valor_mes(bloco_texto, "SAÍDAS (TRANSFERÊNCIA)", idx_mes)
        v_saidas_baixas = extrair_valor_mes(bloco_texto, "SAÍDAS (BAIXAS)", idx_mes)
        
        movim_val = v_dep_mes + v_entradas - v_saidas_transf - v_saidas_baixas
            
        dados_pdf[grupo_id] = {'saldo': saldo_val, 'movimento': movim_val}
            
    return dados_pdf

# O sistema localiza internamente a matriz de cruzamento para relacionar as contas contábeis
def carregar_matriz():
    caminho_matriz = "MATRIZ.xlsx"
    dicionario_matriz = {}
    if os.path.exists(caminho_matriz):
        wb_matriz = openpyxl.load_workbook(caminho_matriz, data_only=True)
        ws_matriz = wb_matriz.active
        for row in ws_matriz.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] is not None:
                dicionario_matriz[str(row[0]).strip()] = str(row[1]).strip()
    return dicionario_matriz

# O sistema gera o layout do documento final em PDF contendo o resultado da sua conciliação
class PDFRelatorio(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 12)
        self.cell(0, 10, 'Relatório de Conciliação - Depreciação Acumulada', 0, 1, 'C')
        self.ln(5)
        
    def footer(self):
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}', 0, 0, 'C')


st.title("📊 Conciliador de Depreciação")

# O sistema recebe o mês que você deseja conciliar para buscar a coluna correta nos documentos
meses_opcoes = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
mes_selecionado = st.selectbox("Selecione o Mês de Referência:", meses_opcoes)
idx_mes = meses_opcoes.index(mes_selecionado)

st.markdown(f"**Mês Base selecionado:** {mes_selecionado}")

with st.expander("📘 GUIA DE USO (Clique para abrir)", expanded=False):
    st.markdown("📌 **Orientações de Uso**")
    st.markdown("""
    1. **Selecione o Mês** que deseja conferir acima.
    2. Anexe a **Planilha do tesouro e os relatórios** todos juntos no mesmo local abaixo.
    3. Clique em "Gerar Relatório".
    """)

# O sistema consolida o recebimento de todos os seus documentos de uma única vez neste botão
uploaded_files = st.file_uploader(
    "Arraste ou selecione a Planilha e os Relatórios de uma só vez", 
    type=['pdf', 'xlsx'], 
    accept_multiple_files=True
)

if st.button("Gerar Relatório de Conciliação", type="primary"):
    if not uploaded_files:
        st.warning("⚠️ Por favor, insira seus arquivos para que possamos realizar a conciliação.")
    else:
        pdfs = [f for f in uploaded_files if f.name.lower().endswith('.pdf')]
        excel_alvo = [f for f in uploaded_files if f.name.lower().endswith('.xlsx')]
        
        if not excel_alvo:
            st.error("❌ A planilha base em Excel não foi encontrada junto com os arquivos enviados.")
            st.stop()
            
        arquivo_alvo = excel_alvo[0]

        with st.spinner("Analisando seus documentos e cruzando os valores..."):
            try:
                dicionario_matriz = carregar_matriz()

                # O sistema lê e armazena os valores de movimentação de todos os PDFs fornecidos
                dados_pdfs_extraidos = {}
                for f in pdfs:
                    uid = extrair_id_unidade(f.name)
                    if uid:
                        dados_pdfs_extraidos[uid] = processar_pdf(f, idx_mes)

                # O sistema percorre sua planilha para agrupar os lançamentos de cada unidade gestora
                wb_alvo = openpyxl.load_workbook(arquivo_alvo, data_only=True)
                
                pdf_out = PDFRelatorio()
                pdf_out.set_auto_page_break(auto=True, margin=15)
                pdf_out.add_page()
                lista_resumo = []
                
                progresso = st.progress(0)
                status_box = st.empty()
                
                abas = [s for s in wb_alvo.sheetnames if s != "MATRIZ"]
                total_abas = len(abas)

                for idx, sheet_name in enumerate(abas):
                    ws = wb_alvo[sheet_name]
                    uid = extrair_id_unidade(sheet_name)
                    
                    status_box.text(f"Conferindo unidade: {sheet_name}...")
                    
                    d_excel = {}
                    
                    for row in ws.iter_rows(values_only=True):
                        if not row or not row[0]: continue
                        
                        conta_raw = str(row[0]).strip()
                        
                        if conta_raw.startswith("12") and conta_raw.replace('.', '').isdigit():
                            if conta_raw == "123110402":
                                continue
                                
                            nat_desp = dicionario_matriz.get(conta_raw)
                            
                            if nat_desp:
                                grupo = extrair_codigo_grupo(nat_desp)
                                if grupo is not None:
                                    valid_vals = [v for v in row if v is not None and str(v).strip() != ""]
                                    if len(valid_vals) >= 2:
                                        saldo_raw = valid_vals[-1]
                                        movim_raw = valid_vals[-2]
                                    elif len(valid_vals) == 1:
                                        saldo_raw = valid_vals[-1]
                                        movim_raw = 0.0
                                    else:
                                        saldo_raw, movim_raw = 0.0, 0.0
                                                
                                    val_saldo = converter_valor_excel(saldo_raw)
                                    val_mov = converter_valor_excel(movim_raw)
                                    
                                    if grupo not in d_excel:
                                        d_excel[grupo] = {'saldo': 0.0, 'movimento': 0.0}
                                        
                                    d_excel[grupo]['saldo'] += val_saldo
                                    d_excel[grupo]['movimento'] += val_mov

                    # O sistema coloca os valores do PDF e do Excel lado a lado para encontrar diferenças exatas
                    if uid in dados_pdfs_extraidos:
                        d_pdf = dados_pdfs_extraidos[uid]
                        
                        grupos = sorted(list(set(d_pdf.keys()) | set(d_excel.keys())))
                        divergencias = []
                        soma_pdf_saldo, soma_excel_saldo = 0.0, 0.0
                        soma_pdf_mov, soma_excel_mov = 0.0, 0.0
                        
                        for g in grupos:
                            vp_saldo = round(-1 * d_pdf.get(g, {}).get('saldo', 0.0), 2)
                            ve_saldo = round(d_excel.get(g, {}).get('saldo', 0.0), 2)
                            
                            vp_mov = round(-1 * d_pdf.get(g, {}).get('movimento', 0.0), 2)
                            ve_mov = round(d_excel.get(g, {}).get('movimento', 0.0), 2)
                            
                            soma_pdf_saldo += vp_saldo
                            soma_excel_saldo += ve_saldo
                            soma_pdf_mov += vp_mov
                            soma_excel_mov += ve_mov
                            
                            dif_saldo = round(vp_saldo - ve_saldo, 2)
                            if abs(dif_saldo) > 0.00: 
                                divergencias.append({'grupo': g, 'tipo': 'Saldo Acumulado', 'pdf': vp_saldo, 'excel': ve_saldo, 'diff': dif_saldo})
                                
                            dif_mov = round(vp_mov - ve_mov, 2)
                            if abs(dif_mov) > 0.00: 
                                divergencias.append({'grupo': g, 'tipo': 'Mês Corrente', 'pdf': vp_mov, 'excel': ve_mov, 'diff': dif_mov})

                        # O sistema desenha a página do relatório consolidado apontando os totais e as possíveis falhas
                        if pdf_out.get_y() > 240: pdf_out.add_page()
                        
                        pdf_out.set_font("helvetica", 'B', 10)
                        pdf_out.set_fill_color(240, 240, 240)
                        pdf_out.cell(0, 8, f"Unidade Gestora: {sheet_name} (ID: {uid})", 1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        pdf_out.ln(2)
                        
                        pdf_out.set_font("helvetica", 'B', 9)
                        pdf_out.set_fill_color(220, 230, 241)
                        pdf_out.cell(48, 7, "Métrica", 1, fill=True, align='C')
                        pdf_out.cell(48, 7, "Total Relatório", 1, fill=True, align='C')
                        pdf_out.cell(48, 7, "Total Planilha", 1, fill=True, align='C')
                        pdf_out.cell(46, 7, "Diferença", 1, fill=True, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        
                        pdf_out.set_font("helvetica", '', 9)
                        pdf_out.cell(48, 7, "Saldo Acumulado", 1)
                        pdf_out.cell(48, 7, f"R$ {formatar_real(soma_pdf_saldo)}", 1, align='R')
                        pdf_out.cell(48, 7, f"R$ {formatar_real(soma_excel_saldo)}", 1, align='R')
                        dif_total_saldo = round(soma_pdf_saldo - soma_excel_saldo, 2)
                        if abs(dif_total_saldo) > 0.00: pdf_out.set_text_color(200, 0, 0)
                        pdf_out.cell(46, 7, f"R$ {formatar_real(dif_total_saldo)}", 1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        pdf_out.set_text_color(0, 0, 0)
                        
                        pdf_out.cell(48, 7, f"Mês Corrente ({mes_selecionado})", 1)
                        pdf_out.cell(48, 7, f"R$ {formatar_real(soma_pdf_mov)}", 1, align='R')
                        pdf_out.cell(48, 7, f"R$ {formatar_real(soma_excel_mov)}", 1, align='R')
                        dif_total_mov = round(soma_pdf_mov - soma_excel_mov, 2)
                        if abs(dif_total_mov) > 0.00: pdf_out.set_text_color(200, 0, 0)
                        pdf_out.cell(46, 7, f"R$ {formatar_real(dif_total_mov)}", 1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        pdf_out.set_text_color(0, 0, 0)
                        
                        pdf_out.ln(3)
                        
                        status_str = "✅ Conciliado"
                        if not divergencias:
                            pdf_out.set_fill_color(220, 255, 220)
                            pdf_out.set_font("helvetica", 'B', 9)
                            pdf_out.cell(0, 8, "CONCILIADO - SEM DIVERGÊNCIAS", 1, fill=True, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        else:
                            status_str = f"❌ Divergência(s)"
                            pdf_out.set_fill_color(255, 220, 220)
                            pdf_out.set_font("helvetica", 'B', 9)
                            pdf_out.cell(0, 8, "DIVERGÊNCIAS ENCONTRADAS:", 1, fill=True, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                            
                            pdf_out.set_fill_color(250, 250, 250)
                            pdf_out.set_font("helvetica", 'B', 8)
                            pdf_out.cell(15, 6, "Grupo", 1, fill=True, align='C')
                            pdf_out.cell(35, 6, "Tipo", 1, fill=True, align='C')
                            pdf_out.cell(46, 6, "Valor Relatório", 1, fill=True, align='C')
                            pdf_out.cell(46, 6, "Valor Planilha", 1, fill=True, align='C')
                            pdf_out.cell(48, 6, "Diferença", 1, fill=True, align='C', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                            
                            pdf_out.set_font("helvetica", '', 8)
                            for d in divergencias:
                                pdf_out.cell(15, 6, str(d['grupo']), 1, align='C')
                                pdf_out.cell(35, 6, d['tipo'], 1, align='C')
                                pdf_out.cell(46, 6, f"R$ {formatar_real(d['pdf'])}", 1, align='R')
                                pdf_out.cell(46, 6, f"R$ {formatar_real(d['excel'])}", 1, align='R')
                                pdf_out.set_text_color(200, 0, 0)
                                pdf_out.cell(48, 6, f"R$ {formatar_real(d['diff'])}", 1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                                pdf_out.set_text_color(0, 0, 0)
                                
                        pdf_out.ln(5)
                        pdf_out.cell(0, 0, "", "B", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                        pdf_out.ln(5)
                        
                        lista_resumo.append({
                            "Unidade / Aba": sheet_name,
                            "Status": status_str,
                            "Dif. Saldo Total": f"R$ {formatar_real(dif_total_saldo)}",
                            "Dif. Mês Corrente": f"R$ {formatar_real(dif_total_mov)}"
                        })
                    else:
                        lista_resumo.append({
                            "Unidade / Aba": sheet_name,
                            "Status": "⚠️ Relatório correspondente não anexado",
                            "Dif. Saldo Total": "-",
                            "Dif. Mês Corrente": "-"
                        })

                    progresso.progress((idx + 1) / total_abas)

                progresso.empty()
                status_box.success("Conferência finalizada com sucesso! Verifique os resultados na tabela e baixe seu relatório final.")
                
                st.markdown("### Resumo Geral da Conciliação")
                st.dataframe(pd.DataFrame(lista_resumo), use_container_width=True)
                
                pdf_bytes = bytes(pdf_out.output()) 
                st.download_button(
                    label="📄 Fazer Download do Relatório Completo (PDF)",
                    data=pdf_bytes,
                    file_name="Relatorio_Conciliacao_Liquida.pdf",
                    mime="application/pdf",
                    type="primary"
                )
                
            except Exception as e:
                st.error(f"Não foi possível processar seus documentos no momento. Verifique se os arquivos estão corretos. (Detalhe: {e})")
